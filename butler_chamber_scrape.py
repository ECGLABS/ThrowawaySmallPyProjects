#!/usr/bin/env python3
"""
Butler County Chamber directory scraper -> XLSX

What it does:
- GET https://www.butlercountychamber.com/directory
- Extract category URLs
- For each category URL, extract business detail-page URLs
- For each business page, extract: Name, Category, Website, Contact, Phone, Email, Address (if present)
- Save to XLSX

Notes:
- Some directory links appear to be broken (404). Script records that as an error row.
- Be a professional: rate limit your requests.
"""

from __future__ import annotations

import argparse
import re
import sys
import time
from dataclasses import dataclass, asdict
from typing import Iterable, Optional
from urllib.parse import urljoin, urlparse

import requests
from bs4 import BeautifulSoup

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


BASE = "https://www.butlercountychamber.com"
DIRECTORY_URL = f"{BASE}/directory"


@dataclass
class BizRow:
    category: str
    business_name: str
    contact: str
    phone: str
    email: str
    website: str
    address: str
    page_url: str
    error: str


def is_internal_url(href: str) -> bool:
    if not href:
        return False
    href = href.strip()
    if href.startswith("#"):
        return False
    if href.startswith("mailto:") or href.startswith("tel:"):
        return False
    if href.startswith("http://") or href.startswith("https://"):
        return urlparse(href).netloc.endswith("butlercountychamber.com")
    return href.startswith("/")


def normalize_url(href: str) -> str:
    href = href.strip()
    if href.startswith("http://") or href.startswith("https://"):
        return href
    return urljoin(BASE, href)


def get_soup(session: requests.Session, url: str, timeout: int = 30) -> BeautifulSoup:
    r = session.get(url, timeout=timeout)
    r.raise_for_status()
    return BeautifulSoup(r.text, "html.parser")


def extract_categories(soup: BeautifulSoup) -> list[tuple[str, str]]:
    """
    Returns list of (category_name, category_url)
    Tries to focus on the actual directory category list by filtering anchor text.
    """
    anchors = soup.find_all("a", href=True)
    cats: list[tuple[str, str]] = []
    seen = set()

    for a in anchors:
        text = " ".join(a.get_text(" ", strip=True).split())
        href = a.get("href", "").strip()
        if not text or not href:
            continue

        # Category labels on this site are usually uppercase words like "ACCOUNTING"
        # Avoid navbar items like "Directory", "Events", etc.
        if len(text) < 3:
            continue
        if text.upper() != text:
            continue

        bad = {"DIRECTORY", "EVENTS", "ENERGY", "TRAVEL", "CONTACT", "JOIN", "BACK", "SKIP TO CONTENT"}
        if text in bad:
            continue

        if not is_internal_url(href):
            continue

        url = normalize_url(href)

        # Heuristic: category pages are typically short slugs like /accounting
        path = urlparse(url).path.strip("/")
        if not path or path == "directory":
            continue

        key = (text, url)
        if key in seen:
            continue
        seen.add(key)
        cats.append((text, url))

    # Stable ordering
    cats.sort(key=lambda x: x[0])
    return cats


def extract_business_links(category_soup: BeautifulSoup) -> list[str]:
    """
    Category pages show businesses as headings with links.
    We'll collect internal links that look like detail pages, and exclude obvious nav links.
    """
    anchors = category_soup.find_all("a", href=True)
    links: list[str] = []
    seen = set()

    for a in anchors:
        href = a.get("href", "").strip()
        text = " ".join(a.get_text(" ", strip=True).split())

        if not is_internal_url(href):
            continue

        url = normalize_url(href)
        path = urlparse(url).path.strip("/")

        # Skip category page itself + obvious site sections
        skip_paths = {
            "", "directory", "events", "energy", "travel", "contact", "join",
            "about", "about/demographics", "leadership", "young-professionals"
        }
        if path in skip_paths:
            continue

        # Skip social-type or doc links
        if "issuu.com" in url:
            continue

        # Must have some text (business name usually)
        if not text:
            continue

        # Don't re-add the same link
        if url in seen:
            continue
        seen.add(url)
        links.append(url)

    return links


LABEL_RE = re.compile(r"^(Contact|Phone|Email|Address)\s*:\s*(.*)\s*$", re.IGNORECASE)


def parse_business_page(soup: BeautifulSoup) -> dict[str, str]:
    """
    Extracts fields from a business page.
    The pages tend to have:
    - A main heading with the business name
    - A "Website" link
    - Plain text lines like "Contact: Name" and "Phone: 724-..."
    """
    # Business name: first h1/h2 in main content
    name = ""
    for tag in soup.find_all(["h1", "h2"]):
        t = tag.get_text(" ", strip=True)
        if t:
            name = t
            break

    website = ""
    for a in soup.find_all("a", href=True):
        t = a.get_text(" ", strip=True).lower()
        if "website" == t:
            website = a["href"].strip()
            break

    # Pull label lines
    contact = phone = email = address = ""
    text_lines = [ln.strip() for ln in soup.get_text("\n", strip=True).splitlines() if ln.strip()]

    for ln in text_lines:
        m = LABEL_RE.match(ln)
        if not m:
            continue
        label = m.group(1).lower()
        val = m.group(2).strip()
        if label == "contact":
            contact = val
        elif label == "phone":
            phone = val
        elif label == "email":
            email = val
        elif label == "address":
            address = val

    # Sometimes email might exist as a mailto link even if no "Email:" line
    if not email:
        mailto = soup.find("a", href=re.compile(r"^mailto:", re.IGNORECASE))
        if mailto and mailto.get("href"):
            email = mailto["href"].split(":", 1)[-1].strip()

    return {
        "business_name": name,
        "website": website,
        "contact": contact,
        "phone": phone,
        "email": email,
        "address": address,
    }


def autosize_columns(ws) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            v = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(v))
        ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)


def write_xlsx(rows: list[BizRow], out_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Butler Chamber Directory"

    headers = list(asdict(rows[0]).keys()) if rows else [
        "category", "business_name", "contact", "phone", "email", "website", "address", "page_url", "error"
    ]
    ws.append(headers)

    for r in rows:
        d = asdict(r)
        ws.append([d[h] for h in headers])

    autosize_columns(ws)
    wb.save(out_path)


def main() -> int:
    categories = extract_categories(dir_soup)
    print(f"[INFO] Found {len(categories)} categories")
    for c, u in categories:
        print(f"  - {c}: {u}")

    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default="butler_chamber_directory.xlsx", help="Output XLSX filename")
    ap.add_argument("--delay", type=float, default=0.8, help="Seconds to sleep between requests")
    ap.add_argument("--timeout", type=int, default=30, help="HTTP timeout seconds")
    args = ap.parse_args()

    session = requests.Session()
    session.headers.update({
        "User-Agent": "Mozilla/5.0 (compatible; DirectoryScraper/1.0; +https://example.com)"
    })

    rows: list[BizRow] = []

    # 1) Directory -> categories
    try:
        dir_soup = get_soup(session, DIRECTORY_URL, timeout=args.timeout)
    except Exception as e:
        print(f"[FATAL] Failed to load directory page: {e}", file=sys.stderr)
        return 2

    categories = extract_categories(dir_soup)
    if not categories:
        print("[FATAL] Found 0 categories. Site structure likely changed.", file=sys.stderr)
        return 3

    # 2) Each category -> business links
    for cat_name, cat_url in categories:
        time.sleep(args.delay)
        try:
            cat_soup = get_soup(session, cat_url, timeout=args.timeout)
        except Exception as e:
            rows.append(BizRow(
                category=cat_name,
                business_name="",
                contact="",
                phone="",
                email="",
                website="",
                address="",
                page_url=cat_url,
                error=f"Category fetch failed: {e}",
            ))
            continue

        biz_links = extract_business_links(cat_soup)
        if not biz_links:
            # Not fatal; just record
            rows.append(BizRow(
                category=cat_name,
                business_name="",
                contact="",
                phone="",
                email="",
                website="",
                address="",
                page_url=cat_url,
                error="No business links found on category page",
            ))
            continue

        # 3) Each business page -> details
        for biz_url in biz_links:
            time.sleep(args.delay)
            try:
                biz_soup = get_soup(session, biz_url, timeout=args.timeout)
                fields = parse_business_page(biz_soup)
                rows.append(BizRow(
                    category=cat_name,
                    business_name=fields.get("business_name", "") or "",
                    contact=fields.get("contact", "") or "",
                    phone=fields.get("phone", "") or "",
                    email=fields.get("email", "") or "",
                    website=fields.get("website", "") or "",
                    address=fields.get("address", "") or "",
                    page_url=biz_url,
                    error="",
                ))
            except requests.HTTPError as e:
                rows.append(BizRow(
                    category=cat_name,
                    business_name="",
                    contact="",
                    phone="",
                    email="",
                    website="",
                    address="",
                    page_url=biz_url,
                    error=f"Business page HTTP error: {e}",
                ))
            except Exception as e:
                rows.append(BizRow(
                    category=cat_name,
                    business_name="",
                    contact="",
                    phone="",
                    email="",
                    website="",
                    address="",
                    page_url=biz_url,
                    error=f"Business parse failed: {e}",
                ))

    write_xlsx(rows, args.out)
    print(f"[OK] Wrote {len(rows)} rows -> {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
